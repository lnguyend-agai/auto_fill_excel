import pandas as pd
import datetime
import openpyxl
from typing import List

def auto_fill(user_list: List[str], file_path: str):
    """
    Auto-fills the specified cells in the Excel file based on the user list and current day.
    """
    new_list = map_user_names(user_list)
    current_day = get_current_day()

    xls = pd.ExcelFile(file_path)
    wb = openpyxl.load_workbook(file_path, data_only=False)
    
    sheet_names_wb = wb.sheetnames
    if sheet_names_wb:
        last_sheet_name = sheet_names_wb[-1]
    
    formula_cells = {}

    last_work_sheet = wb[last_sheet_name]
    formula_cells[last_sheet_name] = {}
    for row in last_work_sheet.iter_rows():
        for cell in row:
            if cell.data_type == 'f':  # Check if the cell contains a formula
                formula_cells[last_sheet_name][cell.coordinate] = cell.value

    # Get the sheet names
    sheet_names = xls.sheet_names

    # Check if there are any sheets in the workbook
    if sheet_names:
        last_sheet_name = sheet_names[-1]
        df = pd.read_excel(file_path, engine='openpyxl', sheet_name=last_sheet_name)

    print("Original Data:")
    print(df)

    df = update_cells(new_list, current_day, df)
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        df.to_excel(writer, sheet_name=last_sheet_name, index=False)

    print("\nUpdated DataFrame:")
    print(df)
    
    # Reapply the formulas
    wb = openpyxl.load_workbook(file_path)
    for sheet_name, cells in formula_cells.items():
        ws = wb[sheet_name]
        for cell_coord, formula in cells.items():
            ws[cell_coord] = formula
    
    wb.save(file_path)

def update_cells(new_list: List[str], current_day: str, df, price: int = 30):
    """
    Updates the cells in the DataFrame for the specified user names and current day.
    """
    for name in new_list:
        if name in df['Ngày'].values:
            df.loc[df['Ngày'] == name, current_day] = price
            
    return df

def get_current_day() -> str:
    """
    Returns the current day of the week in Vietnamese.
    """
    now = datetime.datetime.now()
    day_of_week_name = now.strftime("%A")

    day_mapping = {
        'Monday': 'Thứ 2',
        'Tuesday': 'Thứ 3',
        'Wednesday': 'Thứ 4',
        'Thursday': 'Thứ 5',
        'Friday': 'Thứ 6',
        'Saturday': 'Thứ 7',
        'Sunday': 'Chủ nhật'
    }
    return day_mapping[day_of_week_name]

def map_user_names(user_list: List[str]) -> List[str]:
    """
    Maps the user names from the provided list to their corresponding values.
    """
    name_mapping = {
        'HIEU': 'A.Le.Hiếu',
        'THINH': 'A.Thịnh',
        'TAI': 'A Tài',
        'OKSANA': 'Oksana',
        'CANH': 'Cảnh',
        'BAO': 'Baro',
        'NGOC': 'B.Ngọc',
        'TUYEN': 'Tuyên',
        'NGUYEN': 'Nguyên',
        'THAI': 'Thái',
        'DUC': 'A.Đức',
        'PHAT': 'A Phát Dom',
        'TRUNG': 'A Trung',
        'KHIEM': 'A.Khiêm'  
    }
    return [name_mapping.get(name.upper(), name) for name in user_list]

